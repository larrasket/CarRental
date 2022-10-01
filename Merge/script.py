import os
from openpyxl import Workbook
from pathlib import Path

from openpyxl import load_workbook

home = str(Path.home())
dir_containing_files = home + "/Merge/files/"

dest_wb = Workbook()

for root, dir, filenames in os.walk(dir_containing_files):
    for file in filenames:
        file_name = file.split('.')[0]
        file_path = os.path.abspath(os.path.join(root, file))

        dest_wb.create_sheet(file_name)
        dest_ws = dest_wb[file_name]


        source_wb = load_workbook(file_path)
        source_sheet = source_wb.active
        for row in source_sheet.rows:
            for cell in row:
                dest_ws[cell.coordinate] = cell.value
sv = home + "/Merge/result.xlsx"
dest_wb.save(sv)
